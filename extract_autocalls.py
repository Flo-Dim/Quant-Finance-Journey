"""
Extraction des dates d'observation / remboursement (autocall) depuis des
term sheets PDF, et génération d'un récapitulatif Excel.

Utilisation :
    python extract_autocalls.py
    python extract_autocalls.py --input "C:\\Structurés\\PDF" --output "C:\\Structurés\\autocalls.xlsx"

Dépendances :
    pip install pdfplumber pandas openpyxl
"""

from __future__ import annotations

import argparse
import logging
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import pandas as pd
import pdfplumber
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# =====================================================
# Configuration par défaut (utilisée si --input / --output ne sont pas fournis)
# =====================================================

DEFAULT_INPUT_FOLDER = r"C:\Structurés\PDF"
DEFAULT_OUTPUT_FILE = r"C:\Structurés\autocalls.xlsx"

OBS_KEYWORDS = [
    "observation",
    "evaluation",
    "exercise",
    "valuation",
    "constatation",
]

REMB_KEYWORDS = [
    "remboursement",
    "redemption",
    "paiement",
    "payment",
]

EXCLUDED_TITLE_PATTERNS = [
    "conditions générales",
    "termsheet",
    "term sheet",
    "issuer",
    "émetteur",
]

DATE_PATTERN_FR = re.compile(r"\b(\d{2})/(\d{2})/(\d{4})\b")

DATE_PATTERN_TEXT = re.compile(
    r"\b(January|February|March|April|May|June|July|August|"
    r"September|October|November|December)\s+(\d{1,2})(?:st|nd|rd|th)?,?\s+(\d{4})\b",
    re.IGNORECASE,
)

MONTHS = {
    "january": "01", "february": "02", "march": "03", "april": "04",
    "may": "05", "june": "06", "july": "07", "august": "08",
    "september": "09", "october": "10", "november": "11", "december": "12",
}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("extract_autocalls")


# =====================================================
# Structures de données
# =====================================================

@dataclass
class ProductResult:
    fichier: str
    produit: str
    dates_observation: list[str] = field(default_factory=list)
    dates_remboursement: list[str] = field(default_factory=list)
    statut: str = "OK"


# =====================================================
# Extraction et normalisation des dates
# =====================================================

def _valid_date(day: str, month: str, year: str) -> str | None:
    """Vérifie qu'une date jj/mm/aaaa est réellement valide (pas de 32/13/... etc)."""
    try:
        datetime(int(year), int(month), int(day))
        return f"{day}/{month}/{year}"
    except ValueError:
        return None


def convert_english_date(txt: str) -> str | None:
    m = DATE_PATTERN_TEXT.search(txt)
    if not m:
        return None
    month = MONTHS.get(m.group(1).lower())
    if not month:
        return None
    day = m.group(2).zfill(2)
    year = m.group(3)
    return _valid_date(day, month, year)


def extract_dates_from_text(text: str) -> list[str]:
    if not text:
        return []

    dates: list[str] = []

    for day, month, year in DATE_PATTERN_FR.findall(text):
        valid = _valid_date(day, month, year)
        if valid:
            dates.append(valid)

    for match in DATE_PATTERN_TEXT.finditer(text):
        converted = convert_english_date(match.group(0))
        if converted:
            dates.append(converted)

    return dates


def clean_dates(dates: list[str]) -> list[str]:
    """Déduplique et trie chronologiquement une liste de dates jj/mm/aaaa."""
    unique = set(dates)

    def sort_key(d: str) -> datetime:
        day, month, year = d.split("/")
        return datetime(int(year), int(month), int(day))

    try:
        return sorted(unique, key=sort_key)
    except ValueError:
        # Filet de sécurité : si une date invalide s'est glissée, on la retire.
        cleaned = [d for d in unique if sort_key(d)]
        return sorted(cleaned, key=sort_key)


# =====================================================
# Détection du nom du produit
# =====================================================

def get_product_name(text: str, filename: str) -> str:
    if not text:
        return Path(filename).stem

    for line in text.split("\n")[:20]:
        line = line.strip()
        if 5 < len(line) < 120:
            lower = line.lower()
            if not any(pattern in lower for pattern in EXCLUDED_TITLE_PATTERNS):
                return line

    return Path(filename).stem


# =====================================================
# Extraction depuis les tableaux du PDF
# =====================================================

def extract_from_tables(pdf: pdfplumber.PDF) -> tuple[list[str], list[str]]:
    obs: list[str] = []
    remb: list[str] = []

    for page_num, page in enumerate(pdf.pages, start=1):
        try:
            tables = page.extract_tables()
        except Exception as exc:
            logger.warning("  Page %d : échec extraction tableau (%s)", page_num, exc)
            continue

        if not tables:
            continue

        for table in tables:
            for row in table:
                if not row:
                    continue

                row_text = " ".join(str(cell) for cell in row if cell).lower()

                row_dates: list[str] = []
                for cell in row:
                    if cell:
                        row_dates.extend(extract_dates_from_text(str(cell)))

                if not row_dates:
                    continue

                if any(k in row_text for k in OBS_KEYWORDS):
                    obs.extend(row_dates)
                elif any(k in row_text for k in REMB_KEYWORDS):
                    remb.extend(row_dates)

    return obs, remb


# =====================================================
# Extraction fallback texte brut
# =====================================================

def extract_from_text(pdf: pdfplumber.PDF) -> tuple[list[str], list[str], str]:
    text_parts: list[str] = []

    for page_num, page in enumerate(pdf.pages, start=1):
        try:
            t = page.extract_text()
        except Exception as exc:
            logger.warning("  Page %d : échec extraction texte (%s)", page_num, exc)
            t = None
        if t:
            text_parts.append(t)

    full_text = "\n".join(text_parts)

    obs: list[str] = []
    remb: list[str] = []

    for line in full_text.split("\n"):
        lower = line.lower()
        dates = extract_dates_from_text(line)
        if not dates:
            continue
        if any(k in lower for k in OBS_KEYWORDS):
            obs.extend(dates)
        if any(k in lower for k in REMB_KEYWORDS):
            remb.extend(dates)

    return obs, remb, full_text


# =====================================================
# Traitement d'un fichier PDF
# =====================================================

def process_pdf(pdf_path: Path) -> ProductResult:
    logger.info("Analyse : %s", pdf_path.name)

    try:
        with pdfplumber.open(pdf_path) as pdf:
            if len(pdf.pages) == 0:
                return ProductResult(
                    fichier=pdf_path.name,
                    produit=pdf_path.stem,
                    statut="Erreur : PDF vide (0 page)",
                )

            table_obs, table_remb = extract_from_tables(pdf)
            text_obs, text_remb, raw_text = extract_from_text(pdf)

    except Exception as exc:
        logger.error("  Échec sur %s : %s", pdf_path.name, exc)
        return ProductResult(
            fichier=pdf_path.name,
            produit=pdf_path.stem,
            statut=f"Erreur : {exc}",
        )

    # On garde la source (tableau ou texte brut) qui a trouvé le plus de dates
    obs = table_obs if len(table_obs) >= len(text_obs) else text_obs
    remb = table_remb if len(table_remb) >= len(text_remb) else text_remb

    obs = clean_dates(obs)
    remb = clean_dates(remb)

    product = get_product_name(raw_text, pdf_path.name)

    statut = "OK"
    if not obs and not remb:
        statut = "Attention : aucune date trouvée"

    return ProductResult(
        fichier=pdf_path.name,
        produit=product,
        dates_observation=obs,
        dates_remboursement=remb,
        statut=statut,
    )


# =====================================================
# Génération du fichier Excel
# =====================================================

HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
HEADER_FONT = Font(bold=True)


def build_dataframe(results: list[ProductResult]) -> pd.DataFrame:
    rows = []
    for r in results:
        rows.append({
            "Fichier": r.fichier,
            "Produit": r.produit,
            "Dates Observation": ", ".join(r.dates_observation),
            "Nb Observations": len(r.dates_observation),
            "Dates Remboursement": ", ".join(r.dates_remboursement),
            "Nb Remboursements": len(r.dates_remboursement),
            "Statut": r.statut,
        })
    return pd.DataFrame(rows)


def write_excel(df: pd.DataFrame, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Produits", index=False)

        ws = writer.sheets["Produits"]

        # Mise en forme de l'en-tête : gris, gras, pas de fusion de cellules
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center")

        # Largeur de colonnes ajustée au contenu
        for col_idx, col_name in enumerate(df.columns, start=1):
            max_len = max(
                [len(str(col_name))] + [len(str(v)) for v in df.iloc[:, col_idx - 1]]
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 60)

        ws.freeze_panes = "A2"


# =====================================================
# Point d'entrée principal
# =====================================================

def run(input_folder: Path, output_file: Path) -> int:
    if not input_folder.exists():
        logger.error("Dossier introuvable : %s", input_folder)
        return 1

    pdf_files = sorted(p for p in input_folder.iterdir() if p.suffix.lower() == ".pdf")

    if not pdf_files:
        logger.warning("Aucun fichier PDF trouvé dans %s", input_folder)
        return 0

    logger.info("%d fichier(s) PDF à analyser.", len(pdf_files))

    results: list[ProductResult] = []
    for pdf_path in pdf_files:
        results.append(process_pdf(pdf_path))

    df = build_dataframe(results)
    write_excel(df, output_file)

    n_ok = sum(1 for r in results if r.statut == "OK")
    n_warn = sum(1 for r in results if r.statut.startswith("Attention"))
    n_err = sum(1 for r in results if r.statut.startswith("Erreur"))

    logger.info(
        "Terminé. %d OK, %d sans date, %d en erreur. Excel généré : %s",
        n_ok, n_warn, n_err, output_file,
    )
    return 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input", "-i",
        default=DEFAULT_INPUT_FOLDER,
        help=f"Dossier contenant les PDF (défaut : {DEFAULT_INPUT_FOLDER})",
    )
    parser.add_argument(
        "--output", "-o",
        default=DEFAULT_OUTPUT_FILE,
        help=f"Fichier Excel de sortie (défaut : {DEFAULT_OUTPUT_FILE})",
    )
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    exit_code = run(Path(args.input), Path(args.output))
    sys.exit(exit_code)
